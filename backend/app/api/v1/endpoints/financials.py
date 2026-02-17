from fastapi import APIRouter, Body, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Any
from io import BytesIO
from datetime import datetime
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.config import settings  # se precisar de config/autorização

router = APIRouter(tags=["financials"])

# --- Pydantic models ---
class EntryItem(BaseModel):
    participant_id: Optional[Any] = None
    participant_name: Optional[str] = None
    role: Optional[str] = None
    role_id: Optional[Any] = None
    email: Optional[str] = None
    cpf: Optional[str] = None
    phone: Optional[str] = None
    pix: Optional[str] = None
    bank: Optional[str] = None
    agency: Optional[str] = None
    account: Optional[str] = None
    unit_amount: Optional[float] = 0.0
    days: Optional[int] = 1
    total_per_person: Optional[float] = None
    total_line: Optional[float] = None
    notes: Optional[str] = None

class Options(BaseModel):
    grouped: Optional[bool] = True
    include_bank_details: Optional[bool] = True

class ExportPayload(BaseModel):
    event_id: Optional[int] = None
    event_name: Optional[str] = ""
    entries: List[EntryItem] = Field(default_factory=list)
    options: Optional[Options] = Options()

# --- Helpers ---
def _set_currency(cell):
    cell.number_format = '"R$"#,##0.00'

def _force_text(cell):
    try:
        cell.number_format = '@'
        cell.alignment = Alignment(horizontal="left", vertical="center")
    except Exception:
        pass

def _sanitize_for_excel(value):
    if value is None:
        return ""
    s = str(value)
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s

def format_cpf(raw: Optional[str]) -> str:
    """
    Format CPF to 000.000.000-00 if possible.
    If input doesn't have 11 digits, return original sanitized string.
    """
    if not raw:
        return ""
    digits = re.sub(r'\D', '', str(raw))
    if len(digits) != 11:
        # return original cleaned (but not masked) to avoid losing content
        return re.sub(r'\s+', ' ', str(raw)).strip()
    return f"{digits[0:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:11]}"

def format_phone(raw: Optional[str]) -> str:
    """
    Format Brazilian phone numbers with +55 country code when possible.
    Examples:
     - input 11999998888 -> +55 (11) 99999-8888
     - input 1198888555 -> +55 (11) 9888-5555
     - accepts inputs with or without country code.
    Fallback: return original cleaned string.
    """
    if not raw:
        return ""
    s = str(raw).strip()
    # Remove all non-digits
    digits = re.sub(r'\D', '', s)

    # If input has leading country code '55' or '+55', normalize
    if digits.startswith('55') and len(digits) >= 4:
        body = digits[2:]
    else:
        body = digits

    # Now body should be DDD + number or just number
    if len(body) == 11:
        ddd = body[0:2]
        rest = body[2:]
        return f"+55 ({ddd}) {rest[0:5]}-{rest[5:9]}"
    if len(body) == 10:
        ddd = body[0:2]
        rest = body[2:]
        return f"+55 ({ddd}) {rest[0:4]}-{rest[4:8]}"
    if len(body) == 9:  # no DDD (rare)
        return f"+55 {body[0:5]}-{body[5:9]}"
    if len(body) == 8:
        return f"+55 {body[0:4]}-{body[4:8]}"

    # fallback: return original (trimmed)
    return s

# --- Endpoint: export XLSX ---
@router.post("/financials/export", summary="Gerar planilha financeira (XLSX)")
async def export_financials(payload: ExportPayload = Body(...)):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Financeiro"

        headers = [
            'Evento', 'Participante', 'Função', 'E-mail', 'CPF', 'Telefone',
            'Valor unit. (R$)', 'Dias trabalhados', 'Total por pessoa (R$)',
            'PIX', 'Banco', 'Agência', 'Conta', 'Total da linha (R$)', 'Observações'
        ]

        # Title (merged)
        title = f"Relatório Financeiro — Evento: {payload.event_name or payload.event_id or ''}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(size=14, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Compute total geral
        total_geral = 0.0
        for e in payload.entries:
            line_total = e.total_line if e.total_line is not None else (
                (e.total_per_person if e.total_per_person is not None else (
                    (e.unit_amount or 0.0) * (e.days or 1)
                ))
            )
            total_geral += float(line_total or 0.0)

        # Total geral row (merged)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        c2 = ws.cell(row=2, column=1, value=f"Total Geral: {total_geral:.2f}")
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="left", vertical="center")

        # Spacer row 3; headers at row 4
        header_row = 4
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Data rows start at row 5
        row_idx = header_row + 1
        for e in payload.entries:
            total_per_person = e.total_per_person if e.total_per_person is not None else ((e.unit_amount or 0.0) * (e.days or 1))
            total_line = e.total_line if e.total_line is not None else total_per_person

            # Format CPF and Phone here (ensure masked strings)
            cpf_formatted = format_cpf(e.cpf)
            phone_formatted = format_phone(e.phone)

            values = [
                _sanitize_for_excel(payload.event_name or ""),
                _sanitize_for_excel(e.participant_name or ""),
                _sanitize_for_excel(e.role or ""),
                _sanitize_for_excel(e.email or ""),
                cpf_formatted,
                phone_formatted,
                float(e.unit_amount or 0.0),
                int(e.days or 0),
                float(total_per_person or 0.0),
                _sanitize_for_excel(e.pix or ""),
                _sanitize_for_excel(e.bank or ""),
                _sanitize_for_excel(e.agency or ""),
                _sanitize_for_excel(e.account or ""),
                float(total_line or 0.0),
                _sanitize_for_excel(e.notes or "")
            ]

            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=v)

                # Currency columns: 7,9,14 (1-based)
                if col in (7, 9, 14):
                    try:
                        cell.value = float(v)
                        _set_currency(cell)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    except Exception:
                        pass

                # Days column (8)
                if col == 8:
                    try:
                        cell.value = int(v)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception:
                        pass

                # CPF and Telefone: force text and left align
                if col in (5, 6):
                    try:
                        # write formatted string and force text format
                        cell.value = str(v)
                        _force_text(cell)
                    except Exception:
                        pass

                # borders
                thin = Side(border_style="thin", color="DDDDDD")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            row_idx += 1

        # Bottom total row
        ws.cell(row=row_idx + 1, column=len(headers) - 1, value="Total Geral:")
        total_cell = ws.cell(row=row_idx + 1, column=len(headers), value=total_geral)
        _set_currency(total_cell)
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Column widths
        widths = [30, 28, 20, 28, 16, 18, 15, 14, 16, 18, 14, 10, 14, 18, 25]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header row
        ws.freeze_panes = ws["A5"]

        # Save to bytes and return
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"financeiro_{(payload.event_name or payload.event_id or 'export')}_{datetime.utcnow().date().isoformat()}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
        return StreamingResponse(stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
